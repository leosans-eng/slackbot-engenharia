import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def aplicar_borda_contorno(ws, start_row, start_col, end_row, end_col):
    borda_externa = Side(style="thin", color="000000")
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            if row in (start_row, end_row) or col in (start_col, end_col):
                celula = ws.cell(row=row, column=col)
                celula.border = Border(
                    left=borda_externa if col == start_col else celula.border.left,
                    right=borda_externa if col == end_col else celula.border.right,
                    top=borda_externa if row == start_row else celula.border.top,
                    bottom=borda_externa if row == end_row else celula.border.bottom,
                )


def finalizar_total_secao(ws, row_secao, primeira_linha, ultima_linha, align_direita, linhas_totais_secoes):
    cel_secao_total = ws.cell(row=row_secao, column=6)
    cel_secao_total.alignment = align_direita
    if primeira_linha > ultima_linha:
        cel_secao_total.value = "A ORÇAR"
    else:
        cel_secao_total.value = f"=SUM(F{primeira_linha}:F{ultima_linha})"
        cel_secao_total.number_format = '_-"R$" * #,##0.00_-'
        linhas_totais_secoes.append(row_secao)


def extrair_totais_finais(ws, max_rows=3):
    totais = []
    linha = ws.max_row
    while linha >= 1 and len(totais) < max_rows:
        valor_i = ws.cell(row=linha, column=9).value
        valor_k = ws.cell(row=linha, column=11).value
        if valor_i is not None or valor_k is not None:
            label = str(ws.cell(row=linha, column=1).value or "").strip()
            totais.append({
                "label": label,
                "valor_i": valor_i,
                "valor_k": valor_k,
            })
        linha -= 1
    return list(reversed(totais))


def ajustar_estetica_modelo2(caminho_origem_xlsx, caminho_saida_xlsx):
    # 1. Carregar a planilha de origem (.xlsx)
    try:
        wb_origem = openpyxl.load_workbook(caminho_origem_xlsx)
        ws_origem = wb_origem.active
    except FileNotFoundError:
        print(f"\n🔴 ERRO: O arquivo '{caminho_origem_xlsx}' não foi encontrado.")
        print("Certifique-se de que ele está na mesma pasta que este script e com o nome idêntico.\n")
        return

    # 2. Criar a nova planilha de destino estilizada
    wb_destino = openpyxl.Workbook()
    ws_destino = wb_destino.active
    ws_destino.title = "Orçamento Formatado"
    ws_destino.views.sheetView[0].showGridLines = True
    
    # --- PALETA DE CORES E FONTES ---
    # Usar Aptos Narrow conforme solicitado
    font_cabecalho = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")
    fill_cabecalho = PatternFill(start_color="595959", end_color="595959", fill_type="solid") # Cabeçalho das colunas (segunda linha)
    
    font_secao = Font(name="Aptos Narrow", size=12, bold=True, color="000000")
    fill_secao = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    
    font_corpo = Font(name="Aptos Narrow", size=11, bold=False, color="000000")
    
    align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_esquerda = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_direita = Alignment(horizontal="right", vertical="center")

    # Linha de título principal (acima da tabela)
    top_title = "ORÇAMENTO - REPAROS DE VÍCIOS CONSTRUTIVOS - SINAPI"
    colunas_modelo = [
        "Código", "Descrição", "Un.", "Qtd.", 
        "Preço Unit.", "Total sem BDI"
    ]

    ws_destino.append([top_title] + [""] * (len(colunas_modelo) - 1))
    ws_destino.row_dimensions[1].height = 28
    # Mesclar a primeira linha sobre todas as colunas
    ws_destino.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas_modelo))
    cel_top = ws_destino.cell(row=1, column=1)
    cel_top.font = Font(name="Aptos Narrow", size=12, bold=True, color="FFFFFF")
    cel_top.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    cel_top.alignment = Alignment(horizontal="center", vertical="center")

    # Linha de cabeçalhos de coluna (segunda linha)
    ws_destino.append(colunas_modelo)
    ws_destino.row_dimensions[2].height = 22
    for col_idx, nome_coluna in enumerate(colunas_modelo, 1):
        celula = ws_destino.cell(row=2, column=col_idx)
        celula.font = font_cabecalho
        celula.fill = fill_cabecalho
        celula.alignment = align_centro

    # Encontrar a linha onde começam os dados reais na tabela de origem
    linha_dados_start = 13  
    for row_idx in range(1, ws_origem.max_row + 1):
        val_celula = ws_origem.cell(row=row_idx, column=1).value
        if val_celula and str(val_celula).strip() == "Item":
            linha_dados_start = row_idx + 1
            break
            
    ultima_secao_row = None
    primeira_linha_secao = None
    linhas_totais_secoes = []

    # 3. Mapear e transferir os dados da planilha velha para a nova
    for row_idx in range(linha_dados_start, ws_origem.max_row + 1):
        # Ignora linhas complementares vazias
        item_val = ws_origem.cell(row=row_idx, column=1).value

        # Detectar texto por extenso antes de ignorar linha
        texto_extenso = str(ws_origem.cell(row=row_idx, column=4).value or "").strip()

        if (
            (item_val is None or str(item_val).strip() == "")
            and texto_extenso == ""
        ):
            continue
            
        item = str(item_val).strip()
        banco = str(ws_origem.cell(row=row_idx, column=3).value or "").strip()
        codigo_original = str(ws_origem.cell(row=row_idx, column=4).value or "").strip()

        if banco.lower() == "próprio":
            codigo = "Comp. SINAPI"
        else:
            codigo = codigo_original

        descricao = str(ws_origem.cell(row=row_idx, column=5).value or "").strip()

        # Cálculo automático de altura
        # Quantidade aproximada de caracteres por linha
        limite_linha = 85

        # Quantas linhas o texto ocupará
        num_linhas = max(1, (len(descricao) // limite_linha) + 1)

        # Altura baseada na quantidade de linhas
        altura_linha = num_linhas * 15

        if codigo == "Comp. SINAPI":
            altura_linha = max(altura_linha, 30)  # Garantir altura mínima para itens de seção

        # Leitura direta dos valores do Excel (openpyxl já converte para número automaticamente)
        qtd = ws_origem.cell(row=row_idx, column=7).value
        preco_s_bdi = ws_origem.cell(row=row_idx, column=8).value
        total_s_bdi = ws_origem.cell(row=row_idx, column=10).value
        
        num_linha_atual = ws_destino.max_row + 1
        ws_destino.row_dimensions[num_linha_atual].height = altura_linha
        
        # LINHA DE TEXTO POR EXTENSO
        if (
            texto_extenso
            and banco == ""
            and descricao == ""
        ):
            ws_destino.append(["", "", "", texto_extenso, "", ""])

            linha_extenso = ws_destino.max_row

            # Mesclar B até E
            ws_destino.merge_cells(
                start_row=linha_extenso,
                start_column=2,
                end_row=linha_extenso,
                end_column=5
            )

            celula_extenso = ws_destino.cell(row=linha_extenso, column=2)

            celula_extenso.font = Font(
                name="Aptos Narrow",
                size=9,
                bold=False
            )

            celula_extenso.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )

            # Altura automática
            linhas_extenso = max(1, (len(texto_extenso) // 90) + 1)
            ws_destino.row_dimensions[linha_extenso].height = linhas_extenso * 14

            continue

        # É uma linha de item principal?
        if codigo_original == "":
            # Finalizar seção anterior (sem fórmula se não houver subitens)
            if ultima_secao_row is not None and primeira_linha_secao is not None:
                ultima_linha_secao = num_linha_atual - 1
                finalizar_total_secao(
                    ws_destino, ultima_secao_row, primeira_linha_secao,
                    ultima_linha_secao, align_direita, linhas_totais_secoes,
                )

            ultima_secao_row = num_linha_atual
            primeira_linha_secao = num_linha_atual + 1
            ws_destino.append(["", descricao, "", "", "", 0])
            
            # Aumentar altura para itens principais
            ws_destino.row_dimensions[num_linha_atual].height = max(altura_linha, 26)
            
            # Mesclar B até D
            ws_destino.merge_cells(
                start_row=num_linha_atual,
                start_column=2,
                end_row=num_linha_atual,
                end_column=4
            )
            
            for c_idx in range(1, 7):
                celula = ws_destino.cell(row=num_linha_atual, column=c_idx)
                celula.fill = fill_secao
                # Títulos das seções: fonte em negrito tamanho 12 (font_secao)
                celula.font = font_secao
                if c_idx == 1:
                    celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif c_idx in [2, 3]:
                    celula.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif c_idx == 6:
                    celula.alignment = align_direita
        else:
            # É uma linha de item regular
            un = str(ws_origem.cell(row=row_idx, column=6).value or "").strip()
            ws_destino.append([codigo, descricao, un, qtd, preco_s_bdi, f"=D{num_linha_atual}*E{num_linha_atual}"])
            
            for c_idx in range(1, 7):
                celula = ws_destino.cell(row=num_linha_atual, column=c_idx)
                celula.font = font_corpo
                
                if c_idx == 1:
                    celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif c_idx == 2:
                    celula.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif c_idx == 3:
                    celula.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    celula.alignment = Alignment(horizontal="general", vertical="center")
                    # Colunas monetárias
                    if c_idx in [5, 6]:
                        celula.number_format = '_-"R$" * #,##0.00_-'

    # Finalizar última seção
    if ultima_secao_row is not None and primeira_linha_secao is not None:
        ultima_linha_secao = ws_destino.max_row
        if ultima_secao_row not in linhas_totais_secoes:
            finalizar_total_secao(
                ws_destino, ultima_secao_row, primeira_linha_secao,
                ultima_linha_secao, align_direita, linhas_totais_secoes,
            )

    totais_finais = extrair_totais_finais(ws_origem)
    tabela_final = ws_destino.max_row
    if totais_finais:
        # Linha 1: TOTAL PARCIAL (A:D), soma das linhas principais em F
        linha1 = ws_destino.max_row + 1
        ws_destino.merge_cells(start_row=linha1, start_column=1, end_row=linha1, end_column=4)
        cel = ws_destino.cell(row=linha1, column=1)
        cel.value = "TOTAL PARCIAL="
        cel.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cel.font = Font(name="Aptos Narrow", size=12, bold=True)
        cel.fill = fill_secao

        ws_destino.merge_cells(start_row=linha1, start_column=5, end_row=linha1, end_column=6)
        cel_e = ws_destino.cell(row=linha1, column=5)
        cel_e.font = font_secao
        cel_e.fill = fill_secao
        cel_e.alignment = align_direita
        if linhas_totais_secoes:
            cel_e.value = "=" + "+".join(f"F{row}" for row in linhas_totais_secoes)
        else:
            cel_e.value = 0
        cel_e.number_format = '_-"R$" * #,##0.00_-'

        # Linha 2: BDI - BENEFÍCIOS... (A:C), D = 30,62% com fundo A6A6A6, valor em E:F
        linha2 = linha1 + 1
        ws_destino.merge_cells(start_row=linha2, start_column=1, end_row=linha2, end_column=3)
        cel = ws_destino.cell(row=linha2, column=1)
        cel.value = "BDI - BENEFÍCIOS E DESPESAS INDIRETAS ="
        cel.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cel.font = Font(name="Aptos Narrow", size=12, bold=True)
        cel.fill = fill_secao

        cel_d = ws_destino.cell(row=linha2, column=4)
        cel_d.value = "30,62%"
        cel_d.alignment = align_centro
        cel_d.font = Font(name="Aptos Narrow", size=12, bold=True)
        cel_d.fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

        ws_destino.merge_cells(start_row=linha2, start_column=5, end_row=linha2, end_column=6)
        cel_e = ws_destino.cell(row=linha2, column=5)
        cel_e.font = font_secao
        cel_e.fill = fill_secao
        cel_e.alignment = align_direita
        cel_e.value = f"=E{linha1}*D{linha2}"
        cel_e.number_format = '_-"R$" * #,##0.00_-'

        # Linha 3: ORÇAMENTO TOTAL (A:D), destaque com fonte 14 branca, fundo 595959
        linha3 = linha2 + 1
        ws_destino.merge_cells(start_row=linha3, start_column=1, end_row=linha3, end_column=4)
        cel = ws_destino.cell(row=linha3, column=1)
        cel.value = "ORÇAMENTO TOTAL"
        cel.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cel.font = Font(name="Aptos Narrow", size=14, bold=True, color="FFFFFF")
        cel.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")

        ws_destino.merge_cells(start_row=linha3, start_column=5, end_row=linha3, end_column=6)
        cel_e = ws_destino.cell(row=linha3, column=5)
        cel_e.font = Font(name="Aptos Narrow", size=14, bold=True, color="FFFFFF")
        cel_e.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        cel_e.alignment = align_direita
        cel_e.value = f"=SUM(E{linha1}:F{linha2})"
        cel_e.number_format = '_-"R$" * #,##0.00_-'

        # Aplicar borda externa preta ao bloco de totais (linhas 1..3, colunas 1..6)
        aplicar_borda_contorno(ws_destino, linha1, 1, linha3, 6)

    # 4. Ajuste de Margem/Largura das Colunas
        
    ws_destino.column_dimensions['A'].width = 8
    ws_destino.column_dimensions['B'].width = 80
    ws_destino.column_dimensions['C'].width = 5
    ws_destino.column_dimensions['D'].width = 8
    ws_destino.column_dimensions['E'].width = 14
    ws_destino.column_dimensions['F'].width = 18
    
    dados_final = tabela_final
    # Borda principal começa na linha 2 (após o título superior)
    aplicar_borda_contorno(ws_destino, 2, 1, dados_final, 6)
    wb_destino.save(caminho_saida_xlsx)
    print(f"\n🟢 Sucesso! Planilha gerada no formato do Modelo-2: '{caminho_saida_xlsx}'\n")
    
    # Abrir o arquivo Excel gerado
    caminho_absoluto = os.path.abspath(caminho_saida_xlsx)
    os.startfile(caminho_absoluto)

# --- Configuração de Execução ---
if __name__ == "__main__":
    arquivo_origem_sistema = "Planilha Sintética Simples 1017 .xlsx"
    arquivo_saida_formatado = "Planilha_Sintetica_Convertida_Modelo2.xlsx"
    
    ajustar_estetica_modelo2(arquivo_origem_sistema, arquivo_saida_formatado)
