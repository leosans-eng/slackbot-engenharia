"""Utilitários compartilhados entre os modelos de formatação."""

from __future__ import annotations

import os
import re

from num2words import num2words
from openpyxl.styles import Border, Side

FORMATO_MOEDA = '_-"R$" * #,##0.00_-'


def valor_em_extenso(valor):
    """Converte um valor numérico para texto por extenso em português (formato monetário)."""
    if valor is None or valor == "":
        return ""
    try:
        valor_float = float(valor)
        parte_inteira = int(valor_float)
        parte_centavos = round((valor_float - parte_inteira) * 100)
        texto_inteira = num2words(parte_inteira, lang="pt_BR")
        texto_centavos = num2words(parte_centavos, lang="pt_BR")
        return f"{texto_inteira.capitalize()} reais e {texto_centavos} centavos"
    except (ValueError, TypeError):
        return ""


def extrair_nome_obra(ws_origem, linha=4):
    """Extrai o nome da obra da linha indicada."""
    for col in range(1, ws_origem.max_column + 1):
        valor = ws_origem.cell(row=linha, column=col).value
        if not valor:
            continue
        texto = str(valor).strip()
        match = re.search(r"obra\s*:\s*(.+)", texto, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return ""


def extrair_referencia_sinapi(ws_origem, linha=8):
    """Extrai referência SINAPI (ex.: 'SINAPI PE (04/2026)') da linha de Bancos."""
    for col in range(1, ws_origem.max_column + 1):
        valor = ws_origem.cell(row=linha, column=col).value
        if not valor:
            continue
        texto = str(valor).strip()
        match = re.search(
            r"SINAPI:\s*([A-Z]{2})\s*(\d{1,2})/(\d{4})",
            texto,
            flags=re.IGNORECASE,
        )
        if match:
            uf = match.group(1).upper()
            mes = match.group(2).zfill(2)
            ano = match.group(3)
            return f"SINAPI {uf} ({mes}/{ano})"
    return ""


def sanitizar_nome_arquivo(nome):
    """Remove caracteres inválidos para nomes de arquivo no Windows."""
    nome_limpo = re.sub(r'[<>:"/\\|?*]', "-", nome)
    return re.sub(r"\s+", " ", nome_limpo).strip()


def gerar_caminho_saida(
    caminho_origem,
    numero_modelo,
    nome_obra="",
    *,
    caminho_saida=None,
    diretorio_saida=None,
):
    """Gera o caminho do Excel formatado."""
    if caminho_saida:
        return os.path.abspath(caminho_saida)

    diretorio = diretorio_saida or os.path.dirname(os.path.abspath(caminho_origem)) or "."
    if nome_obra:
        nome_limpo = sanitizar_nome_arquivo(nome_obra)
        nome_arquivo = (
            f"Planilha_Sintetica_Convertida_Modelo{numero_modelo} - {nome_limpo}.xlsx"
        )
    else:
        nome_arquivo = f"Planilha_Sintetica_Convertida_Modelo{numero_modelo}.xlsx"
    return os.path.join(diretorio, nome_arquivo)


def aplicar_borda_contorno(ws, start_row, start_col, end_row, end_col):
    borda_externa = Side(style="thin", color="000000")
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            if row in (start_row, end_row) or col in (start_col, end_col):
                celula = ws.cell(row=row, column=col)
                celula.border = Border(
                    left=borda_externa if col == start_col else celula.border.left,
                    right=borda_externa if col == end_col else celula.border.right,
                    top=borda_externa if row == start_row else celula.border.top,
                    bottom=borda_externa if row == end_row else celula.border.bottom,
                )


def extrair_totais_finais(ws, max_rows=3):
    totais = []
    linha = ws.max_row
    while linha >= 1 and len(totais) < max_rows:
        valor_i = ws.cell(row=linha, column=9).value
        valor_k = ws.cell(row=linha, column=11).value
        if valor_i is not None or valor_k is not None:
            totais.append({
                "label": str(ws.cell(row=linha, column=1).value or "").strip(),
                "valor_i": valor_i,
                "valor_k": valor_k,
            })
        linha -= 1
    return list(reversed(totais))


def encontrar_linha_dados_start(ws_origem, padrao=13):
    for row_idx in range(1, ws_origem.max_row + 1):
        val_celula = ws_origem.cell(row=row_idx, column=1).value
        if val_celula and str(val_celula).strip() == "Item":
            return row_idx + 1
    return padrao
