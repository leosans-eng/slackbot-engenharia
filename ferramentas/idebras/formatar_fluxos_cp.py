"""Formata Excel de fluxos do CP no padrão modelo/fluxos-cp.xlsx."""

from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

HEADERS = [
    "Pasta/ID",
    "Nome",
    "Processo",
    "Condomínio",
    "Cidade",
    "Prazo Eng.",
    "Prazo Fatal",
    "Observação",
]

COL_WIDTHS = {
    "A": 10.14,
    "B": 9.86,
    "C": 12.0,
    "D": 9.86,
    "E": 7.14,
    "F": 15.43,
    "G": 16.14,
    "H": 38.86,
    "I": 9.14,
}

FONT_HEADER = Font(name="Calibri", size=11, bold=True, italic=True)
FONT_BODY = Font(name="Calibri", size=11, bold=False, italic=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
FORMATO_DATA = "DD/MM/YYYY"
FORMATO_ID = "0"

HEADER_ALIASES: dict[str, int] = {
    "pasta/id": 0,
    "nome": 1,
    "nomecliente": 1,
    "cliente": 1,
    "processo": 2,
    "nº processo": 2,
    "n° processo": 2,
    "numeroprocesso": 2,
    "condominio": 3,
    "conjunto": 3,
    "conjuntohabitacional": 3,
    "cidade": 4,
    "prazo eng.": 5,
    "prazo eng": 5,
    "prazo engenharia": 5,
    "prazo escritorio": 5,
    "data prazo escritorio": 5,
    "prazo fatal": 6,
    "data prazo fatal": 6,
    "fatal": 6,
    "observacao": 7,
    "observacoes": 7,
}


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    for a, b in (
        ("á", "a"),
        ("à", "a"),
        ("â", "a"),
        ("ã", "a"),
        ("é", "e"),
        ("ê", "e"),
        ("í", "i"),
        ("ó", "o"),
        ("ô", "o"),
        ("õ", "o"),
        ("ú", "u"),
        ("ç", "c"),
    ):
        text = text.replace(a, b)
    return re.sub(r"\s+", " ", text)


def _as_datetime(value: Any) -> datetime | date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, time):
        return None
    text = str(value).strip()
    for fmt in (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%m-%d-%y",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _find_header_row(ws: Worksheet) -> tuple[int, dict[int, int]]:
    """Retorna (linha_header, mapa col_origem_1based -> índice_modelo)."""
    for row in range(1, min(ws.max_row, 40) + 1):
        mapping: dict[int, int] = {}
        used_dest: set[int] = set()
        cells = [
            (col, _norm_header(ws.cell(row, col).value))
            for col in range(1, ws.max_column + 1)
        ]
        for col, key in cells:
            if key in {"pasta/id", "data prazo fatal", "data prazo escritorio"}:
                dest = HEADER_ALIASES[key]
                mapping[col] = dest
                used_dest.add(dest)
        for col, key in cells:
            if not key or key not in HEADER_ALIASES:
                continue
            dest = HEADER_ALIASES[key]
            if dest in used_dest:
                continue
            if key == "trabalho a executar" and 7 in used_dest:
                continue
            mapping[col] = dest
            used_dest.add(dest)

        dests = set(mapping.values())
        if {0, 1, 2}.issubset(dests):
            return row, mapping
        if {0, 1, 2, 3, 4}.issubset(dests):
            return row, mapping
    raise RuntimeError(
        "Não encontrou cabeçalho do Excel de fluxos CP "
        "(esperava Pasta/ID, Nome, Processo, …)."
    )


def _iter_rows(ws: Worksheet) -> list[list[Any]]:
    header_row, mapping = _find_header_row(ws)
    rows: list[list[Any]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        values: list[Any] = [None] * 8
        empty = True
        for src_col, dest in mapping.items():
            val = ws.cell(row, src_col).value
            if val is not None and str(val).strip() != "":
                empty = False
            if dest < 8:
                if dest == 7 and values[7] is not None and values[7] != "":
                    continue
                values[dest] = val
        if empty:
            continue
        if isinstance(values[0], str) and values[0].strip().lower() in {"total", "soma"}:
            continue
        rows.append(values)
    return rows


def formatar_fluxos_cp(origem: Path | str, destino: Path | str) -> Path:
    """Lê Excel bruto do CP e grava no padrão do modelo."""
    origem = Path(origem)
    destino = Path(destino)

    wb_raw = load_workbook(origem, data_only=False)
    ws_raw = wb_raw.active
    assert ws_raw is not None
    raw_rows = _iter_rows(ws_raw)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Planilha1"

    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    for col, text in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col, text)
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER

    for offset, raw in enumerate(raw_rows):
        excel_row = offset + 2
        pasta_id, nome, processo, condo, cidade, _prazo_eng, _prazo_fatal, obs = raw

        values = [pasta_id, nome, processo, condo, cidade, None, None, obs]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(excel_row, col, value)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_CENTER if col in {1, 5, 6, 7} else ALIGN_LEFT

        id_cell = ws.cell(excel_row, 1)
        if isinstance(id_cell.value, (int, float)):
            id_cell.number_format = FORMATO_ID

        eng = ws.cell(excel_row, 6)
        eng.value = f"=G{excel_row}-3"
        eng.number_format = FORMATO_DATA
        eng.font = FONT_BODY
        eng.alignment = ALIGN_CENTER

        fatal = ws.cell(excel_row, 7)
        fatal.number_format = FORMATO_DATA
        fatal.alignment = ALIGN_CENTER

    last_row = max(1, len(raw_rows) + 1)
    ws.auto_filter.ref = f"A1:H{last_row}"

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino
