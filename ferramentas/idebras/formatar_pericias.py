"""Formata o Excel bruto de Perícia Finalizada no padrão do relatório."""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PrintPageSetup
from openpyxl.worksheet.worksheet import Worksheet

SHEET_NAME = "Perícia Finalizada"
TITLE = "RELATÓRIO DE LAUDOS DE IMPUGNAÇÃO FINALIZADOS"

FILL_TITLE = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
FILL_HEADER = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
FONT_TITLE = Font(name="Calibri", size=24, bold=True, color="000000")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_HEADER_PRAZO = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11, bold=False, color="000000")
FONT_TIPO = Font(name="Calibri", size=10, bold=False, color="000000")
FONT_TOTAL = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
SIDE_MEDIUM = Side(style="medium", color="000000")
BORDER_TITLE_LEFT = Border(
    left=SIDE_MEDIUM, right=SIDE_MEDIUM, top=SIDE_MEDIUM, bottom=SIDE_MEDIUM
)
BORDER_TITLE_MID = Border(top=SIDE_MEDIUM, bottom=SIDE_MEDIUM)
BORDER_TITLE_RIGHT = Border(
    right=SIDE_MEDIUM, top=SIDE_MEDIUM, bottom=SIDE_MEDIUM
)
BORDER_ROW = Border(bottom=Side(style="thin", color="C0C0C0"))

PRINT_MARGIN = 0.07874015748031496
PRINT_PAPER_A4 = 9

FORMATO_MOEDA = '_-"R$" * #,##0.00_-;\\-"R$" * #,##0.00_-;_-"R$" * "-"??_-;_-@_-'
FORMATO_PCT = "0.00%"
FORMATO_DATA = "DD/MM/YYYY"

HEADERS = [
    "ID",
    "Nº Processo",
    "Autor",
    "Valor Perito",
    "Valor Engenharia",
    "Diferenca R$",
    "Resultado",
    "Tipo",
    "Prazo Peticionamento",
]

COL_WIDTHS = {
    "A": 7.14,
    "B": 24.43,
    "C": 40.0,
    "D": 16.29,
    "E": 14.71,
    "F": 9.43,
    "G": 13.14,
    "H": 10.0,
    "I": 11.0,
    "J": 9.14,
}

RAW_COL = {
    "id": 2,
    "processo": 6,
    "autor": 9,
    "valor_perito": 18,
    "valor_engenharia": 20,
    "diferenca_pct": 22,
    "resultado": 23,
    "prazo": 24,
}


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _as_datetime(value: Any) -> datetime | date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, time):
        return None
    text = str(value).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d", "%m-%d-%y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _find_header_row(ws: Worksheet) -> int:
    for row in range(1, min(ws.max_row, 30) + 1):
        values = [str(ws.cell(row, c).value or "").strip().lower() for c in range(1, 12)]
        if "nº processo" in values or "n° processo" in values or "cod. processo" in values:
            if any("autor" in v for v in values) or any("valor perito" in v for v in values):
                return row
        if "cód. processo" in values or "cod. processo" in values:
            if "valor perito" in values or "resultado" in values:
                return row
    raise RuntimeError(
        "Não encontrou a linha de cabeçalho no Excel bruto de perícias."
    )


def _iter_raw_rows(ws: Worksheet) -> list[list[Any]]:
    header_row = _find_header_row(ws)
    rows: list[list[Any]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]
        first = values[0]
        if first is None and all(v is None for v in values):
            continue
        if isinstance(first, str) and first.strip().lower() == "total":
            continue
        codigo = values[RAW_COL["id"]] if len(values) > RAW_COL["id"] else None
        processo = values[RAW_COL["processo"]] if len(values) > RAW_COL["processo"] else None
        if codigo is None and processo is None:
            continue
        rows.append(values)
    return rows


def _apply_column_widths(ws: Worksheet) -> None:
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width


def _apply_title_outer_border(ws: Worksheet) -> None:
    for col in range(2, 8):
        cell = ws.cell(1, col)
        cell.fill = FILL_TITLE
        if col == 2:
            cell.border = BORDER_TITLE_LEFT
        elif col == 7:
            cell.border = BORDER_TITLE_RIGHT
        else:
            cell.border = BORDER_TITLE_MID


def _write_title_block(ws: Worksheet, emissao: datetime) -> None:
    ws.merge_cells("B1:G1")
    title = ws["B1"]
    title.value = TITLE
    title.font = FONT_TITLE
    title.fill = FILL_TITLE
    title.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32.25
    _apply_title_outer_border(ws)

    label = ws["B2"]
    label.value = "Data Emissão"
    label.font = FONT_HEADER
    label.fill = FILL_HEADER
    label.alignment = ALIGN_CENTER

    data = ws["B3"]
    data.value = emissao.strftime("%d/%m/%Y %H:%M:%S")
    data.font = FONT_BODY
    data.alignment = ALIGN_CENTER


def _apply_print_layout(ws: Worksheet, last_row: int) -> None:
    ws.print_area = f"A1:I{max(last_row, 4)}"
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = PRINT_MARGIN
    ws.page_margins.right = PRINT_MARGIN
    ws.page_margins.top = PRINT_MARGIN
    ws.page_margins.bottom = PRINT_MARGIN
    ws.page_margins.header = PRINT_MARGIN
    ws.page_margins.footer = PRINT_MARGIN

    ws.page_setup = PrintPageSetup(
        worksheet=ws,
        orientation="landscape",
        paperSize=PRINT_PAPER_A4,
        scale=99,
    )
    ws.sheet_view.view = "pageBreakPreview"
    ws.sheet_view.zoomScale = 85
    ws.sheet_view.zoomScaleNormal = 85
    ws.sheet_view.showGridLines = False


def _write_headers(ws: Worksheet) -> None:
    ws.row_dimensions[4].height = 35.25
    for idx, text in enumerate(HEADERS, start=1):
        cell = ws.cell(4, idx, text)
        cell.font = FONT_HEADER_PRAZO if idx == 9 else FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER_WRAP
    ws.auto_filter.ref = "A4:I4"


def _tipo_formula(row: int) -> str:
    return (
        f'=IF(OR(G{row}="IMPUGNAR",G{row}="MANIFESTAR"),"-",'
        f'IF(F{row}>-30%,"TOTAL","PARCIAL"))'
    )


def _write_data_row(ws: Worksheet, excel_row: int, raw: list[Any]) -> None:
    id_val = raw[RAW_COL["id"]]
    processo = raw[RAW_COL["processo"]]
    autor = raw[RAW_COL["autor"]]
    valor_perito = _as_float(raw[RAW_COL["valor_perito"]])
    valor_eng = _as_float(raw[RAW_COL["valor_engenharia"]])
    diff_pct = _as_float(raw[RAW_COL["diferenca_pct"]])
    resultado = raw[RAW_COL["resultado"]]
    prazo = _as_datetime(raw[RAW_COL["prazo"]])

    values = [
        id_val,
        processo,
        autor,
        valor_perito,
        valor_eng,
        diff_pct,
        resultado,
        _tipo_formula(excel_row),
        prazo.date() if isinstance(prazo, datetime) else prazo,
    ]

    for col, value in enumerate(values, start=1):
        cell = ws.cell(excel_row, col, value)
        cell.font = FONT_TIPO if col == 8 else FONT_BODY
        cell.alignment = ALIGN_CENTER if col == 8 else ALIGN_LEFT
        if col == 3 and isinstance(autor, str) and len(autor) > 28:
            cell.alignment = ALIGN_LEFT_WRAP
        cell.border = BORDER_ROW if col != 8 else Border()
        if col in (4, 5) and value is not None:
            cell.number_format = FORMATO_MOEDA
        elif col == 6 and value is not None:
            cell.number_format = FORMATO_PCT
        elif col == 9 and value is not None:
            cell.number_format = FORMATO_DATA


def _write_total_row(ws: Worksheet, first_data: int, last_data: int) -> None:
    total_row = last_data + 1
    for col in range(1, 10):
        cell = ws.cell(total_row, col)
        cell.fill = FILL_HEADER
        cell.font = FONT_TOTAL
        cell.alignment = Alignment(vertical="center")

    if last_data >= first_data:
        ws.cell(total_row, 4).value = f"=SUM(D{first_data}:D{last_data})"
        ws.cell(total_row, 4).number_format = FORMATO_MOEDA
        ws.cell(total_row, 5).value = f"=SUM(E{first_data}:E{last_data})"
        ws.cell(total_row, 5).number_format = FORMATO_MOEDA
        ws.cell(total_row, 6).value = (
            f"=IF(E{total_row}=0,0,(D{total_row}-E{total_row})/E{total_row})"
        )
        ws.cell(total_row, 6).number_format = FORMATO_PCT


def formatar_pericias_finalizadas(
    origem: Path | str,
    destino: Path | str,
    *,
    data_emissao: datetime | None = None,
) -> Path:
    """Lê o Excel bruto do sistema e grava o relatório formatado."""
    origem = Path(origem)
    destino = Path(destino)
    emissao = data_emissao or datetime.now()

    wb_raw = load_workbook(origem, data_only=False)
    if SHEET_NAME in wb_raw.sheetnames:
        ws_raw = wb_raw[SHEET_NAME]
    else:
        ws_raw = wb_raw.active
    raw_rows = _iter_raw_rows(ws_raw)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = SHEET_NAME

    _apply_column_widths(ws)
    _write_title_block(ws, emissao)
    _write_headers(ws)

    first_data = 5
    for offset, raw in enumerate(raw_rows):
        _write_data_row(ws, first_data + offset, raw)
    last_data = first_data + len(raw_rows) - 1 if raw_rows else first_data - 1
    if raw_rows:
        _write_total_row(ws, first_data, last_data)
        last_row = last_data + 1
    else:
        last_row = 4

    _apply_print_layout(ws, last_row)

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino


def formatar_bytes_pericias(
    data: bytes,
    destino: Path | str,
    *,
    data_emissao: datetime | None = None,
) -> Path:
    """Formata a partir dos bytes baixados (grava temporário interno)."""
    import tempfile

    destino = Path(destino)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)
    try:
        return formatar_pericias_finalizadas(
            tmp_path, destino, data_emissao=data_emissao
        )
    finally:
        tmp_path.unlink(missing_ok=True)
