import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from num2words import num2words

def valor_em_extenso(valor):
    """Converte um valor numérico para texto por extenso em português (formato monetário)."""
    if valor is None or valor == "":
        return ""
    try:
        valor_float = float(valor)
        
        # Separar parte inteira e decimal
        parte_inteira = int(valor_float)
        parte_centavos = round((valor_float - parte_inteira) * 100)
        
        # Converter parte inteira para extenso
        texto_inteira = num2words(parte_inteira, lang='pt_BR')
        
        # Converter parte de centavos para extenso
        texto_centavos = num2words(parte_centavos, lang='pt_BR')
        
        # Montar o texto completo
        resultado = f"{texto_inteira.capitalize()} reais e {texto_centavos} centavos"
        
        return resultado
    except (ValueError, TypeError):
        return ""

def extrair_nome_obra(ws_origem, linha=4):
    """Extrai o nome da obra da linha indicada (ex.: 'Obra: ALTO DO MOURA XI - CARUARU / PE')."""
    for col in range(1, ws_origem.max_column + 1):
        valor = ws_origem.cell(row=linha, column=col).value
        if not valor:
            continue
        texto = str(valor).strip()
        match = re.search(r"obra\s*:\s*(.+)", texto, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return ""


def sanitizar_nome_arquivo(nome):
    """Remove caracteres inválidos para nomes de arquivo no Windows."""
    nome_limpo = re.sub(r'[<>:"/\\|?*]', "-", nome)
    return re.sub(r"\s+", " ", nome_limpo).strip()


def gerar_caminho_saida_modelo1(caminho_origem, nome_obra):
    diretorio = os.path.dirname(os.path.abspath(caminho_origem)) or "."
    if nome_obra:
        nome_limpo = sanitizar_nome_arquivo(nome_obra)
        nome_arquivo = f"Planilha_Sintetica_Convertida_Modelo1 - {nome_limpo}.xlsx"
    else:
        nome_arquivo = "Planilha_Sintetica_Convertida_Modelo1.xlsx"
    return os.path.join(diretorio, nome_arquivo)


def aplicar_borda_contorno(ws, start_row, start_col, end_row, end_col):
    borda_externa = Side(style="thin", color="000000")
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            if row in (start_row, end_row) or col in (start_col, end_col):
                celula = ws.cell(row=row, column=col)
                celula.border = Border(
                    left=borda_externa if col == start_col else celula.border.left,
                    right=borda_externa if col == end_col else celula.border.right,
                    top=borda_externa if row == start_row else celula.border.top,
                    bottom=borda_externa if row == end_row else celula.border.bottom,
                )


def extrair_totais_finais(ws, max_rows=3):
    totais = []
    linha = ws.max_row
    while linha >= 1 and len(totais) < max_rows:
        valor_i = ws.cell(row=linha, column=9).value
        valor_k = ws.cell(row=linha, column=11).value
        if valor_i is not None or valor_k is not None:
            label = str(ws.cell(row=linha, column=1).value or "").strip()
            totais.append({
                "label": label,
                "valor_i": valor_i,
                "valor_k": valor_k,
            })
        linha -= 1
    return list(reversed(totais))


def ajustar_estetica_modelo1(caminho_origem_xlsx):
    # 1. Carregar a planilha de origem (.xlsx)
    try:
        wb_origem = openpyxl.load_workbook(caminho_origem_xlsx)
        ws_origem = wb_origem.active
    except FileNotFoundError:
        print(f"\n🔴 ERRO: O arquivo '{caminho_origem_xlsx}' não foi encontrado.")
        print("Certifique-se de que ele está na mesma pasta que este script e com o nome idêntico.\n")
        return

    nome_obra = extrair_nome_obra(ws_origem)
    caminho_saida_xlsx = gerar_caminho_saida_modelo1(caminho_origem_xlsx, nome_obra)

    # 2. Criar a nova planilha de destino estilizada
    wb_destino = openpyxl.Workbook()
    ws_destino = wb_destino.active
    ws_destino.title = "Orçamento Formatado"
    ws_destino.views.sheetView[0].showGridLines = True
    
    # --- PALETA DE CORES E FONTES ---
    font_cabecalho = Font(name="Calibri", size=11, bold=True, color="000000")
    fill_cabecalho = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid") # Verde
    
    font_secao = Font(name="Calibri", size=12, bold=False, color="000000")
    fill_secao = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    
    font_corpo = Font(name="Calibri", size=10, bold=False, color="000000")
    
    align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_esquerda = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_direita = Alignment(horizontal="right", vertical="center")
    
    borda_fina = Side(style='thin', color='EDEDED')
    grade_celula = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    formato_moeda = '_-"R$" * #,##0.00_-'
    
    # Cabeçalho do Modelo
    colunas_modelo = [
        "Item", "Código Sinapi", "Descrição", "Un.", "Qtd.", 
        "Preço s/ BDI", "Preço c/ BDI", "Total s/ BDI", "Total c/ BDI"
    ]
    ws_destino.append(colunas_modelo)
    ws_destino.row_dimensions[1].height = 32
    
    for col_idx, nome_coluna in enumerate(colunas_modelo, 1):
        celula = ws_destino.cell(row=1, column=col_idx)
        celula.font = font_cabecalho
        celula.fill = fill_cabecalho
        celula.alignment = align_centro
        celula.border = grade_celula

    # Encontrar a linha onde começam os dados reais na tabela de origem
    linha_dados_start = 13  
    for row_idx in range(1, ws_origem.max_row + 1):
        val_celula = ws_origem.cell(row=row_idx, column=1).value
        if val_celula and str(val_celula).strip() == "Item":
            linha_dados_start = row_idx + 1
            break
            
    # 3. Mapear e transferir os dados da planilha velha para a nova
    for row_idx in range(linha_dados_start, ws_origem.max_row + 1):
        # Ignora linhas complementares vazias
        item_val = ws_origem.cell(row=row_idx, column=1).value

        # Detectar texto por extenso antes de ignorar linha
        texto_extenso = str(ws_origem.cell(row=row_idx, column=4).value or "").strip()

        if (
            (item_val is None or str(item_val).strip() == "")
            and texto_extenso == ""
        ):
            continue
            
        item = str(item_val).strip()
        banco = str(ws_origem.cell(row=row_idx, column=3).value or "").strip()
        codigo_original = str(ws_origem.cell(row=row_idx, column=4).value or "").strip()

        if banco.lower() == "próprio":
            codigo = "Comp. SINAPI"
        else:
            codigo = codigo_original

        descricao = str(ws_origem.cell(row=row_idx, column=5).value or "").strip()

        # Cálculo automático de altura
        # Quantidade aproximada de caracteres por linha
        limite_linha = 85

        # Quantas linhas o texto ocupará
        num_linhas = max(1, (len(descricao) // limite_linha) + 1)

        # Altura baseada na quantidade de linhas
        altura_linha = num_linhas * 15

        if codigo == "Comp. SINAPI":
            altura_linha = max(altura_linha, 30)  # Garantir altura mínima para itens de seção

        # Leitura direta dos valores do Excel (openpyxl já converte para número automaticamente)
        qtd = ws_origem.cell(row=row_idx, column=7).value
        preco_s_bdi = ws_origem.cell(row=row_idx, column=8).value
        preco_c_bdi = ws_origem.cell(row=row_idx, column=9).value
        total_s_bdi = ws_origem.cell(row=row_idx, column=10).value
        total_c_bdi = ws_origem.cell(row=row_idx, column=11).value
        
        num_linha_atual = ws_destino.max_row + 1
        ws_destino.row_dimensions[num_linha_atual].height = altura_linha
        
        # LINHA DE TEXTO POR EXTENSO
        if (
            texto_extenso
            and banco == ""
            and descricao == ""
        ):
            ws_destino.append(["", "", "", texto_extenso, "", "", "", "", ""])

            linha_extenso = ws_destino.max_row

            # Mesclar D até H
            ws_destino.merge_cells(
                start_row=linha_extenso,
                start_column=4,
                end_row=linha_extenso,
                end_column=8
            )

            celula_extenso = ws_destino.cell(row=linha_extenso, column=4)

            celula_extenso.font = Font(
                name="Calibri",
                size=9,
                bold=False
            )

            celula_extenso.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )

            # Altura automática
            linhas_extenso = max(1, (len(texto_extenso) // 90) + 1)
            ws_destino.row_dimensions[linha_extenso].height = linhas_extenso * 14

            continue

        # É uma linha de item principal?
        if codigo_original == "":
            valor_extenso = valor_em_extenso(total_c_bdi)
            ws_destino.append([item, "", descricao, valor_extenso, "", "", "", "", total_c_bdi])
            
            # Aumentar altura para itens principais
            ws_destino.row_dimensions[num_linha_atual].height = max(altura_linha, 26)
            
            # Mesclar D até H
            ws_destino.merge_cells(
                start_row=num_linha_atual,
                start_column=4,
                end_row=num_linha_atual,
                end_column=8
            )
            
            for c_idx in range(1, 10):
                celula = ws_destino.cell(row=num_linha_atual, column=c_idx)
                celula.fill = fill_secao
                celula.border = grade_celula
                if c_idx == 4:
                    celula.font = Font(name="Calibri", size=9, bold=False, color="000000")
                else:
                    celula.font = font_secao
                if c_idx in [1, 2]:
                    celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif c_idx == 3:
                    celula.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif c_idx == 4:
                    celula.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif c_idx == 9:
                    celula.alignment = align_direita
                    celula.number_format = formato_moeda
        else:
            # É uma linha de item regular
            un = str(ws_origem.cell(row=row_idx, column=6).value or "").strip()
            ws_destino.append([item, codigo, descricao, un, qtd, preco_s_bdi, preco_c_bdi, total_s_bdi, total_c_bdi])
            
            for c_idx in range(1, 10):
                celula = ws_destino.cell(row=num_linha_atual, column=c_idx)
                celula.font = font_corpo
                celula.border = grade_celula
                
                if c_idx in [1, 2]:
                    celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif c_idx == 3:
                    celula.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif c_idx == 4:
                    celula.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif c_idx == 9:
                    celula.alignment = align_direita
                    celula.number_format = formato_moeda
                else:
                    celula.alignment = Alignment(horizontal="general", vertical="center")
                    if c_idx in [6, 7, 8]:
                        celula.number_format = formato_moeda

    totais_finais = extrair_totais_finais(ws_origem)
    tabela_final = ws_destino.max_row
    if totais_finais:
        labels_default = ["Total sem BDI", "Total do BDI (30,62%)", "Total do Orçamento"]
        for index, tot in enumerate(totais_finais):
            row = tabela_final + 1 + index
            if index == 0:
                label = labels_default[0]
            elif index == 1:
                label = labels_default[1]
            else:
                label = labels_default[2]
            valor = tot["valor_k"] if tot["valor_k"] is not None else tot["valor_i"]
            ws_destino.cell(row=row, column=11).value = label
            ws_destino.cell(row=row, column=12).value = valor

            cel_label = ws_destino.cell(row=row, column=11)
            cel_label.font = font_secao
            cel_label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cel_label.border = grade_celula

            cel_val = ws_destino.cell(row=row, column=12)
            cel_val.font = Font(name="Calibri", size=12, bold=False, color="000000")
            cel_val.alignment = align_direita
            cel_val.border = grade_celula
            if isinstance(valor, (int, float)):
                cel_val.number_format = formato_moeda

            if index == 2:
                cel_label.font = Font(name="Calibri", size=14, bold=True, color="1F2610")
                cel_label.fill = fill_secao
                cel_val.font = Font(name="Calibri", size=14, bold=True, color="1F2610")
                cel_val.fill = fill_secao

        linha_totais_inicio = tabela_final + 1
        linha_totais_fim = tabela_final + len(totais_finais)
        aplicar_borda_contorno(ws_destino, linha_totais_inicio, 11, linha_totais_fim, 12)

        valor_total_orcamento = totais_finais[2]["valor_k"] if len(totais_finais) > 2 and totais_finais[2]["valor_k"] is not None else totais_finais[2]["valor_i"]
        texto_extenso_total = valor_em_extenso(valor_total_orcamento)
        
        linha_extenso_total = linha_totais_fim + 2
        ws_destino.cell(row=linha_extenso_total, column=11).value = texto_extenso_total
        
        cel_extenso = ws_destino.cell(row=linha_extenso_total, column=11)
        cel_extenso.font = Font(name="Calibri", size=10, bold=False, color="000000")
        cel_extenso.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cel_extenso.border = grade_celula

    # 4. Ajuste de Margem/Largura das Colunas
        
    ws_destino.column_dimensions['A'].width = 5
    ws_destino.column_dimensions['B'].width = 8
    ws_destino.column_dimensions['C'].width = 80
    ws_destino.column_dimensions['D'].width = 5
    ws_destino.column_dimensions['E'].width = 7
    ws_destino.column_dimensions['F'].width = 12
    ws_destino.column_dimensions['G'].width = 12
    ws_destino.column_dimensions['H'].width = 14
    ws_destino.column_dimensions['I'].width = 18
    ws_destino.column_dimensions['K'].width = 25
    ws_destino.column_dimensions['L'].width = 22
    
    dados_final = tabela_final
    aplicar_borda_contorno(ws_destino, 1, 1, dados_final, 9)
    wb_destino.save(caminho_saida_xlsx)
    print(f"\n🟢 Sucesso! Planilha gerada no formato do Modelo-1: '{caminho_saida_xlsx}'\n")
    
    # Abrir o arquivo Excel gerado
    caminho_absoluto = os.path.abspath(caminho_saida_xlsx)
    os.startfile(caminho_absoluto)

# --- Configuração de Execução ---
if __name__ == "__main__":
    arquivo_origem_sistema = "Planilha Sintética Simples 1017 .xlsx"
    
    ajustar_estetica_modelo1(arquivo_origem_sistema)